"""
colleges/services.py — Business logic for the College Outreach module.

Sits between the router (HTTP) and the store (persistence). Holds:
  • dashboard aggregation       (Feature 5)
  • outreach status transitions  (Feature 4)
  • CSV / Excel import & export   (Feature 6)
  • a re-score-all helper that fans the prioritiser across the dataset (Feature 7)

Kept free of FastAPI types so it stays unit-testable and reusable.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date
from typing import Optional

from colleges import store
from colleges.prioritizer import apply_score
from colleges.schemas import (
    CollegeRecord, OutreachStatus, CollegeType, PriorityLevel, _now_iso,
)

logger = logging.getLogger("volt_cv.colleges")


# ──────────────────────────────────────────────
# Feature 4 — Outreach status tracking
# ──────────────────────────────────────────────

def update_status(
    college_id: str,
    new_status: OutreachStatus,
    *,
    set_contact_date: bool = True,
    note: str = "",
) -> Optional[CollegeRecord]:
    """Advance a college's outreach status, optionally stamping the contact date
    and appending a timestamped note to the record's notes field."""
    college = store.get(college_id)
    if college is None:
        return None

    patch: dict = {"outreach_status": new_status.value}

    # Stamp contact date when we actively reach out (not for terminal/neutral states)
    contacting = new_status in (
        OutreachStatus.EMAIL_SENT, OutreachStatus.CALL_SCHEDULED,
        OutreachStatus.PARTNERSHIP_DISCUSS,
    )
    if set_contact_date and contacting:
        patch["last_contact_date"] = date.today().isoformat()

    if note:
        stamp = _now_iso()
        existing_notes = college.notes or ""
        patch["notes"] = (existing_notes + f"\n[{stamp}] {new_status.value}: {note}").strip()

    return store.update(college_id, patch)


# ──────────────────────────────────────────────
# Feature 5 — Dashboard aggregation
# ──────────────────────────────────────────────

def dashboard_stats() -> dict:
    """Return JSON shaped for dashboard cards + charts."""
    colleges = store.list_all()
    total = len(colleges)

    # Count by status
    status_counts = {s.value: 0 for s in OutreachStatus.ordered()}
    for c in colleges:
        status_counts[c.outreach_status] = status_counts.get(c.outreach_status, 0) + 1

    # Count by type
    type_counts = {t.value: 0 for t in CollegeType}
    for c in colleges:
        type_counts[c.college_type] = type_counts.get(c.college_type, 0) + 1

    # Count by priority level (compute live so it's always current)
    level_counts = {l.value: 0 for l in PriorityLevel}
    for c in colleges:
        _, level = (c.priority_score, c.priority_level)
        if level is None:
            scored = apply_score(c)
            level = scored.priority_level
        level_counts[level] = level_counts.get(level, 0) + 1

    contacted = total - status_counts[OutreachStatus.NOT_CONTACTED.value]
    awaiting  = status_counts[OutreachStatus.AWAITING_RESPONSE.value]
    interested = (
        status_counts[OutreachStatus.INTERESTED.value]
        + status_counts[OutreachStatus.NEED_MORE_INFO.value]
        + status_counts[OutreachStatus.CALL_SCHEDULED.value]
    )
    active = status_counts[OutreachStatus.ACTIVE_PARTNER.value]

    # Funnel ordered for a chart (drop the terminal NOT_INTERESTED out of the line)
    funnel = [
        {"stage": s.value, "count": status_counts[s.value]}
        for s in OutreachStatus.ordered()
        if s != OutreachStatus.NOT_INTERESTED
    ]

    return {
        "cards": {
            "total_colleges":     total,
            "colleges_contacted": contacted,
            "awaiting_response":  awaiting,
            "interested":         interested,
            "active_partners":    active,
        },
        "by_status":   [{"status": k, "count": v} for k, v in status_counts.items()],
        "by_type":     [{"type": k, "count": v} for k, v in type_counts.items()],
        "by_priority": [{"level": k, "count": v} for k, v in level_counts.items()],
        "funnel":      funnel,
        "conversion_rate": round(active / contacted * 100, 1) if contacted else 0.0,
    }


# ──────────────────────────────────────────────
# Feature 7 — re-score the whole dataset
# ──────────────────────────────────────────────

def rescore_all() -> int:
    """Recompute priority_score/level for every college and persist. Returns count."""
    colleges = store.list_all()
    for c in colleges:
        apply_score(c)
    store.replace_all(colleges)
    logger.info("Re-scored %d colleges", len(colleges))
    return len(colleges)


def top_priorities(limit: int = 10) -> list[dict]:
    """Return the highest-scoring colleges (live-scored) for a 'who to contact next' widget."""
    colleges = store.list_all()
    for c in colleges:
        if c.priority_score is None:
            apply_score(c)
    ranked = sorted(colleges, key=lambda c: c.priority_score or 0, reverse=True)
    return [
        {
            "id": c.id,
            "college_name": c.college_name,
            "priority_score": c.priority_score,
            "priority_level": c.priority_level,
            "outreach_status": c.outreach_status,
            "college_type": c.college_type,
        }
        for c in ranked[:limit]
    ]


# ──────────────────────────────────────────────
# Feature 6 — Import / Export
# ──────────────────────────────────────────────

# Columns used for both CSV and Excel I/O — order is the display order.
EXPORT_COLUMNS = [
    "id", "college_name", "placement_officer_name", "designation",
    "email", "phone", "city", "state", "college_type",
    "website", "placement_page_url", "last_contact_date", "outreach_status",
    "engineering_intake", "internship_opportunities", "placement_quality_score",
    "historical_engagement", "priority_score", "priority_level", "notes",
    "created_at", "updated_at",
]

# Aliases let messy import files map onto our fields (lowercased, stripped).
_COLUMN_ALIASES = {
    "name": "college_name", "college": "college_name", "institution": "college_name",
    "tpo": "placement_officer_name", "placement officer": "placement_officer_name",
    "officer": "placement_officer_name", "contact name": "placement_officer_name",
    "title": "designation", "role": "designation",
    "mail": "email", "email id": "email", "e-mail": "email",
    "mobile": "phone", "contact": "phone", "phone number": "phone",
    "type": "college_type", "category": "college_type",
    "url": "website", "site": "website",
    "placement url": "placement_page_url", "placement page": "placement_page_url",
    "status": "outreach_status",
    "intake": "engineering_intake",
}

_VALID_FIELDS = set(EXPORT_COLUMNS) | {"outreach_status"}


def _coerce_row(raw: dict) -> CollegeRecord:
    """Map a raw import dict (arbitrary headers) onto a CollegeRecord."""
    mapped: dict = {}
    for k, v in raw.items():
        if k is None:
            continue
        key = str(k).strip().lower()
        field = _COLUMN_ALIASES.get(key, key.replace(" ", "_"))
        if field in _VALID_FIELDS and field not in ("id", "created_at", "updated_at"):
            mapped[field] = (str(v).strip() if v is not None else "")

    # Type coercion for numeric fields
    for num in ("engineering_intake", "internship_opportunities",
                "placement_quality_score", "historical_engagement"):
        if num in mapped and mapped[num] != "":
            try:
                mapped[num] = int(float(mapped[num]))
            except (ValueError, TypeError):
                mapped[num] = None

    # Normalise college_type to a valid enum value (best-effort)
    if "college_type" in mapped:
        ct = mapped["college_type"].strip().upper()
        match = next((t.value for t in CollegeType if t.value.upper() == ct), None)
        mapped["college_type"] = match or CollegeType.OTHER.value

    # Default/repair status
    if "outreach_status" in mapped:
        valid = {s.value for s in OutreachStatus.ordered()}
        if mapped["outreach_status"] not in valid:
            mapped["outreach_status"] = OutreachStatus.NOT_CONTACTED.value

    rec = CollegeRecord(**{k: v for k, v in mapped.items() if v not in (None, "")
                           or k == "college_name"})
    return apply_score(rec)


def import_csv(content: bytes, *, on_duplicate: str = "skip") -> dict:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    records = [_coerce_row(row) for row in reader if any(row.values())]
    records = [r for r in records if r.college_name.strip()]
    summary = store.bulk_upsert(records, on_duplicate=on_duplicate)
    summary["parsed"] = len(records)
    return summary


def import_excel(content: bytes, *, on_duplicate: str = "skip") -> dict:
    import openpyxl   # lazy — only needed on Excel import
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "skipped": 0, "parsed": 0}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        raw = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if any(v not in (None, "") for v in raw.values()):
            rec = _coerce_row(raw)
            if rec.college_name.strip():
                records.append(rec)
    summary = store.bulk_upsert(records, on_duplicate=on_duplicate)
    summary["parsed"] = len(records)
    return summary


def export_csv() -> bytes:
    colleges = store.list_all()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for c in colleges:
        writer.writerow(c.to_dict())
    return buf.getvalue().encode("utf-8-sig")


def export_excel() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colleges"

    # Header row with light styling
    header_fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font

    for c in store.list_all():
        d = c.to_dict()
        ws.append([d.get(col, "") for col in EXPORT_COLUMNS])

    # Reasonable column widths
    for col_idx, name in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(14, len(name) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
