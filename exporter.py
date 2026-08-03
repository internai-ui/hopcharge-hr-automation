"""
exporter.py — Writes parsed candidate records to JSON, CSV, and XLSX.

JSON preserves the full nested schema (arrays, sub-objects).
CSV/XLSX flatten arrays to semicolon-delimited strings so every
candidate is one row.
"""

import json
import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

from config import OUTPUT_DIR
from schemas import CandidateRecord

logger = logging.getLogger(__name__)


def _flatten_record(rec: dict) -> dict:
    """Flatten nested lists/dicts into single-value strings for tabular output."""
    flat = {}
    for key, val in rec.items():
        if key == "raw_text":
            flat[key] = val[:500] + "…" if len(val) > 500 else val  # truncate for spreadsheet
        elif key == "field_confidence":
            flat[key] = json.dumps(val)
        elif key == "personal_details" and isinstance(val, dict):
            for sub_k, sub_v in val.items():
                flat[f"personal_{sub_k}"] = sub_v
        elif isinstance(val, list):
            if val and isinstance(val[0], dict):
                # Structured list (work_experience, education)
                parts = []
                for item in val:
                    parts.append(" | ".join(f"{k}: {v}" for k, v in item.items() if v))
                flat[key] = "; ".join(parts)
            else:
                flat[key] = "; ".join(str(v) for v in val)
        else:
            flat[key] = val
    return flat


def export_json(records: list[CandidateRecord], filename: str = "candidates.json") -> Path:
    out_path = OUTPUT_DIR / filename
    data = [r.to_dict() for r in records]
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    logger.info("JSON exported → %s (%d records)", out_path, len(records))
    return out_path


def export_csv(records: list[CandidateRecord], filename: str = "candidates.csv") -> Path:
    out_path = OUTPUT_DIR / filename
    flat = [_flatten_record(r.to_dict()) for r in records]
    df = pd.DataFrame(flat)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    logger.info("CSV exported → %s (%d records)", out_path, len(records))
    return out_path


def export_xlsx(records: list[CandidateRecord], filename: str = "candidates.xlsx") -> Path:
    out_path = OUTPUT_DIR / filename
    flat = [_flatten_record(r.to_dict()) for r in records]
    df = pd.DataFrame(flat)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidates")
        ws = writer.sheets["Candidates"]
        # Style the header row
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        body_font = Font(name="Arial", size=10)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # Auto-width (capped at 50)
            max_len = max(len(str(col_name)), df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) else 0)
            ws.column_dimensions[cell.column_letter].width = min(max_len + 4, 50)
        # Body font
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Freeze header
        ws.freeze_panes = "A2"

    logger.info("XLSX exported → %s (%d records)", out_path, len(records))
    return out_path


def export_all(records: list[CandidateRecord]) -> dict[str, Path]:
    """Export to all three formats and return a dict of paths."""
    return {
        "json": export_json(records),
        "csv": export_csv(records),
        "xlsx": export_xlsx(records),
    }
